from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _q2(v: Any) -> Decimal:
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"))
    try:
        return Decimal(str(v or 0)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _brl_number_format() -> str:
    return '[$R$-pt-BR] #,##0.00'


def _style_header(ws, row: int, col_start: int, col_end: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

    ws.row_dimensions[row].height = 22


def _apply_table_style(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="F7F7F7")

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if r % 2 == 0:
                cell.fill = zebra


def _autosize_columns(ws, max_width: int = 60) -> None:
    dims: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

    for col, w in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, w + 2), max_width)


def _fmt_competencia(comp: str) -> str:
    comp = str(comp or "").strip()
    if len(comp) == 6 and comp.isdigit():
        return f"01-{comp[4:6]}-{comp[:4]}"
    return comp or "000000"


def criar_aba_descricao(wb, linhas_cruzadas):
    from collections import defaultdict
    from decimal import Decimal

    ws = wb.create_sheet("DESCRICAO")

    headers = [
        "COMPETÊNCIA",
        "VALOR ITEM",
        "VALOR ICMS",
        "VALOR IPI",
        "ATUAL - VL BASE PIS",
        "ATUAL - VL PIS",
        "ORIGINAL - VL BASE PIS",
        "ORIGINAL - VL PIS",
        "ATUAL - VL BASE COFINS",
        "ATUAL - VL COFINS",
        "ORIGINAL - VL BASE COFINS",
        "ORIGINAL - VL COFINS",
        "ÊXITO PIS",
        "ÊXITO COFINS",
        "ÊXITO TOTAL",
    ]

    ws.append(headers)
    _style_header(ws, 1, 1, len(headers))

    agg = defaultdict(lambda: {
        "valor_item": Decimal("0"),
        "valor_icms": Decimal("0"),
        "valor_ipi": Decimal("0"),
        "base_pis_atual": Decimal("0"),
        "pis_atual": Decimal("0"),
        "base_pis_orig": Decimal("0"),
        "pis_orig": Decimal("0"),
        "base_cof_atual": Decimal("0"),
        "cof_atual": Decimal("0"),
        "base_cof_orig": Decimal("0"),
        "cof_orig": Decimal("0"),
    })

    for row in linhas_cruzadas:
        if not row.get("dominio_ok"):
            continue

        comp = str(row.get("competencia") or "").strip() or "000000"
        g = agg[comp]

        g["valor_item"] += Decimal(str(row.get("valor_item", 0) or 0))
        g["valor_icms"] += Decimal(str(row.get("valor_icms", 0) or 0))
        g["valor_ipi"] += Decimal(str(row.get("valor_ipi", 0) or 0))

        g["base_pis_atual"] += Decimal(str(row.get("base_simulada", 0) or 0))
        g["pis_atual"] += Decimal(str(row.get("pis_simulado", 0) or 0))

        g["base_cof_atual"] += Decimal(str(row.get("base_simulada", 0) or 0))
        g["cof_atual"] += Decimal(str(row.get("cofins_simulado", 0) or 0))

        g["base_pis_orig"] += Decimal(str(row.get("vl_base_pis", 0) or 0))
        g["pis_orig"] += Decimal(str(row.get("vl_pis", 0) or 0))

        g["base_cof_orig"] += Decimal(str(row.get("vl_base_cofins", 0) or 0))
        g["cof_orig"] += Decimal(str(row.get("vl_cofins", 0) or 0))

    for comp, v in sorted(agg.items()):
        exito_pis = v["pis_atual"] - v["pis_orig"]
        exito_cof = v["cof_atual"] - v["cof_orig"]
        exito_total = exito_pis + exito_cof

        ws.append([
            _fmt_competencia(comp),
            float(v["valor_item"]),
            float(v["valor_icms"]),
            float(v["valor_ipi"]),
            float(v["base_pis_atual"]),
            float(v["pis_atual"]),
            float(v["base_pis_orig"]),
            float(v["pis_orig"]),
            float(v["base_cof_atual"]),
            float(v["cof_atual"]),
            float(v["base_cof_orig"]),
            float(v["cof_orig"]),
            float(exito_pis),
            float(exito_cof),
            float(exito_total),
        ])

    for row in range(2, ws.max_row + 1):
        for col in range(2, 16):
            ws.cell(row, col).number_format = _brl_number_format()

    _apply_table_style(ws, 1, ws.max_row, 1, len(headers))
    _autosize_columns(ws, max_width=28)


def _write_rows(ws, rows: List[Dict[str, Any]], columns: List[str], money_cols: List[str]) -> None:
    ws.append(columns)
    _style_header(ws, 1, 1, len(columns))

    money_idx = {columns.index(col) + 1 for col in money_cols if col in columns}

    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    for r in range(2, ws.max_row + 1):
        for c in money_idx:
            ws.cell(r, c).number_format = _brl_number_format()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
    _apply_table_style(ws, 1, ws.max_row, 1, len(columns))
    _autosize_columns(ws, max_width=80)


def exportar_cruzamento_xlsx(
    out_xlsx: Path,
    *,
    linhas_cruzadas: List[Dict[str, Any]],
    resumo: Dict[str, Any],
) -> None:
    wb = Workbook()

    # ======================================================
    # ABA 1 - RESUMO
    # ======================================================
    ws = wb.active
    ws.title = "RESUMO"

    title_font = Font(bold=True, size=16, color="1F4E79")
    label_font = Font(bold=True, color="404040")

    ws["A1"] = "FOTO RECUPERAÇÃO — CRUZAMENTO ICMS/IPI x EFD CONTRIBUIÇÕES"
    ws["A1"].font = title_font

    resumo_rows = [
        ("Total linhas", resumo.get("total_linhas", 0)),
        ("Total match", resumo.get("total_match", 0)),
        ("Total sem match", resumo.get("total_sem_match", 0)),
        ("Total não escriturado", resumo.get("total_nao_escriturado", 0)),
        ("Total elegíveis", resumo.get("total_elegiveis", 0)),
        ("Base simulada (R$)", resumo.get("total_base_simulada", 0)),
        ("PIS simulado (R$)", resumo.get("total_pis_simulado", 0)),
        ("COFINS simulado (R$)", resumo.get("total_cofins_simulado", 0)),
        ("Crédito simulado total (R$)", resumo.get("total_credito_simulado", 0)),
        ("Crédito com match (R$)", resumo.get("total_credito_match", 0)),
        ("Crédito não escriturado (R$)", resumo.get("total_credito_nao_escriturado", 0)),
    ]

    row = 3
    for label, value in resumo_rows:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=float(_q2(value)) if isinstance(value, Decimal) else value)
        if "R$" in label:
            ws.cell(row=row, column=2).number_format = _brl_number_format()
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Resumo por tipo de match").font = label_font
    row += 1
    ws.append(["Tipo Match", "Quantidade"])
    _style_header(ws, row, 1, 2)

    for k, v in sorted((resumo.get("por_match") or {}).items()):
        ws.append([k, v])

    _apply_table_style(ws, row, ws.max_row, 1, 2)
    _autosize_columns(ws)

    # ======================================================
    # FILTROS
    # ======================================================
    recuperaveis = [x for x in linhas_cruzadas if _q2(x.get("credito_simulado")) > 0]
    nao_escriturados = [x for x in linhas_cruzadas if x.get("status_cruzamento") == "NAO_ESCRITURADO"]
    sem_match = nao_escriturados  # compatibilidade
    todos = linhas_cruzadas

    columns = [
        "competencia",
        "empresa",
        "participante",
        "data",
        "chave",
        "numero",
        "serie",
        "num_item",
        "cod_item",
        "cod_item_norm",
        "descricao",
        "ncm",
        "cfop",
        "valor_item",
        "valor_desconto",
        "valor_icms",
        "valor_ipi",
        "cst_pis",
        "vl_base_pis",
        "vl_aliq_pis",
        "vl_pis",
        "cst_cofins",
        "vl_base_cofins",
        "vl_aliq_cofins",
        "vl_cofins",
        "base_simulada",
        "pis_simulado",
        "cofins_simulado",
        "credito_simulado",
        "cst_simulado",
        "regra_simulada",
        "elegivel_simulacao",
        "motivo_simulacao",
        "dominio_ok",
        "tipo_match",
        "match_encontrado",
        "status_cruzamento",
        "origem",
        "origem_item",
    ]

    money_cols = [
        "valor_item",
        "valor_desconto",
        "valor_icms",
        "valor_ipi",
        "vl_base_pis",
        "vl_aliq_pis",
        "vl_pis",
        "vl_base_cofins",
        "vl_aliq_cofins",
        "vl_cofins",
        "base_simulada",
        "pis_simulado",
        "cofins_simulado",
        "credito_simulado",
    ]

    criar_aba_descricao(wb, linhas_cruzadas)

    # ======================================================
    # ABA 2 - RECUPERAVEIS
    # ======================================================
    ws2 = wb.create_sheet("RECUPERAVEIS")
    _write_rows(ws2, recuperaveis, columns, money_cols)

    # ======================================================
    # ABA 3 - TODOS
    # ======================================================
    ws3 = wb.create_sheet("TODOS")
    _write_rows(ws3, todos, columns, money_cols)

    # ======================================================
    # ABA 4 - NAO_ESCRITURADOS
    # ======================================================
    ws4 = wb.create_sheet("NAO_ESCRITURADOS")
    _write_rows(ws4, nao_escriturados, columns, money_cols)

    # ======================================================
    # ABA 5 - SEM_MATCH (compatibilidade)
    # ======================================================
    ws5 = wb.create_sheet("SEM_MATCH")
    _write_rows(ws5, sem_match, columns, money_cols)

    wb.save(out_xlsx)
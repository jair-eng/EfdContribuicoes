
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

r"""
SCRIPT — FOTO DE OPORTUNIDADE (CAFÉ)
Atualização (2026-02):
- Mantém TODAS as travas de segurança (CPF/PF, participante desconhecido, C170 sem C100).
- Mantém filtro de CFOP (guard-rail) e filtro opcional de NCM.
- Passa a somar a BASE de cálculo como:
    (A) base_credito_existente: itens já creditáveis (CST 50-56)
  + (B) base_credito_geravel: itens que virarão crédito com as alterações implementadas (CST origem, dentro do CFOP-alvo)
- Crédito final é calculado sobre base_total_credito = existente + gerável.
"""

ALIQUOTA_PIS = Decimal("0.0165")
ALIQUOTA_COF = Decimal("0.0760")
ALIQUOTA_TOTAL = (ALIQUOTA_PIS + ALIQUOTA_COF).quantize(Decimal("0.0001"))

# CSTs já creditáveis
CST_CREDITO = {"50", "51", "52", "53", "54", "55", "56"}

# Guard-rail CFOP (café torrado / compra para comercialização)
CFOPS_TORRADO_PADRAO = {"1101", "1102", "2101", "2102", "3101", "3102"}

# CSTs de origem que o motor converte para crédito (ex.: 51) — mantenha aqui
CSTS_ORIGEM_PADRAO = {"70", "73","74", "75", "98", "99", "06", "07", "08"}

NCMS_CAFE_VERDE_PADRAO = set()

CPF_RE = re.compile(r"^\d{11}$")


# --- FUNÇÕES DE APOIO ---

def only_digits(x: object) -> str:
    return "".join(ch for ch in str(x or "") if ch.isdigit())


def norm_cpf(raw: object) -> str:
    d = only_digits(raw)
    if not d:
        return ""
    if len(d) > 11:
        d = d[-11:]
    return d.zfill(11)



def extrair_cnpj_cpf_0150(fields: List[str]) -> Tuple[str, str]:
    """
    Tenta extrair CNPJ/CPF do registro 0150 de forma robusta.

    Layout comum: |0150|COD_PART|NOME|COD_PAIS|CNPJ|CPF|IE|...
    Mas alguns arquivos vêm com colunas deslocadas; então:
      - Prioriza posições 3 (CNPJ) e 4 (CPF)
      - Se não bater (14/11 dígitos), procura em todos os campos um token com 14 dígitos (CNPJ)
        e um token com 11 dígitos (CPF).
    """
    cnpj = fields[3] if len(fields) > 3 else ""
    cpf = fields[4] if len(fields) > 4 else ""

    d_cnpj = only_digits(cnpj)
    d_cpf = only_digits(cpf)

    if len(d_cnpj) != 14:
        for tok in fields:
            if len(only_digits(tok)) == 14:
                cnpj = tok
                d_cnpj = only_digits(tok)
                break

    if len(d_cpf) != 11:
        for tok in fields:
            if len(only_digits(tok)) == 11:
                cpf = tok
                d_cpf = only_digits(tok)
                break

    return cnpj, cpf

def eh_pf_rigoroso(cnpj_raw: str, cpf_raw: str) -> bool:
    """
    True = PF (CPF) / False = PJ (CNPJ) ou ambíguo.
    Regra: se existir QUALQUER CNPJ, NUNCA bloqueia como PF.
    """
    if only_digits(cnpj_raw):
        return False

    cpf = norm_cpf(cpf_raw)
    return len(cpf) == 11 and cpf != "00000000000"


# --- FUNÇÕES DE FORMATAÇÃO E EXCEL ---

def _brl_number_format() -> str:
    return '[$R$-pt-BR] #,##0.00'


def _style_header(ws, row: int, col_start: int, col_end: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, align, border
    ws.row_dimensions[row].height = 22


def _apply_table_style(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra = PatternFill("solid", fgColor="F7F7F7")
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if r % 2 == 0:
                cell.fill = zebra


def _autosize_columns(ws, max_width: int = 60) -> None:
    dims: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))
    for col, w in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, w + 2), max_width)


def salvar_xlsx_visual(
    out_xlsx: Path,
    *,
    pasta: Path,
    ncms_validos: set[str],
    results: List["FotoArquivoResult"],
    total_base_credito_existente: Decimal,
    total_base_credito_geravel: Decimal,
    total_base_total_credito: Decimal,
    total_cred: Decimal,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    title_font = Font(bold=True, size=16, color="1F4E79")
    label_font = Font(bold=True, color="404040")

    ws["A1"] = "FOTO RECUPERÁVEL — CAFÉ (SIMULAÇÃO)"
    ws["A1"].font = title_font

    ws["A3"], ws["B3"] = "Pasta analisada:", str(pasta)
    ws["A3"].font = label_font

    ws["A4"], ws["B4"] = "Filtro NCM:", ("; ".join(sorted(ncms_validos)) if ncms_validos else "Sem filtro")
    ws["A4"].font = label_font

    ws["A6"], ws["B6"] = "Base crédito existente (R$):", float(q2(total_base_credito_existente))
    ws["B6"].number_format = _brl_number_format()
    ws["A6"].font = label_font

    ws["A7"], ws["B7"] = "Base crédito gerável (R$):", float(q2(total_base_credito_geravel))
    ws["B7"].number_format = _brl_number_format()
    ws["A7"].font = label_font

    ws["A8"], ws["B8"] = "Base total crédito (R$):", float(q2(total_base_total_credito))
    ws["B8"].number_format = _brl_number_format()
    ws["A8"].font = label_font

    ws["A9"], ws["B9"] = "Total crédito (R$):", float(q2(total_cred))
    ws["B9"].number_format = _brl_number_format()
    ws["A9"].font = label_font

    ws.append([])
    ws.append(["Período", "Base existente (R$)", "Base gerável (R$)", "Base total (R$)", "Crédito (R$)", "Arquivos"])
    header_row = ws.max_row
    _style_header(ws, header_row, 1, 6)

    agg: Dict[str, Dict[str, Decimal | int]] = {}
    for r in results:
        per = r.periodo_mmyyyy or "??????"
        if r.credito_total <= 0:
            continue
        if per not in agg:
            agg[per] = {
                "base_exist": Decimal("0"),
                "base_ger": Decimal("0"),
                "base_total": Decimal("0"),
                "cred": Decimal("0"),
                "qtd": 0,
            }
        agg[per]["base_exist"] += r.base_credito_existente
        agg[per]["base_ger"] += r.base_credito_geravel
        agg[per]["base_total"] += r.base_total_credito
        agg[per]["cred"] += r.credito_total
        agg[per]["qtd"] += 1

    for per in sorted(agg.keys(), key=mmYYYY_to_key):
        ws.append([
            per,
            float(q2(agg[per]["base_exist"])),
            float(q2(agg[per]["base_ger"])),
            float(q2(agg[per]["base_total"])),
            float(q2(agg[per]["cred"])),
            int(agg[per]["qtd"]),
        ])

    for r in range(header_row + 1, ws.max_row + 1):
        for col in (2, 3, 4, 5):
            ws.cell(r, col).number_format = _brl_number_format()

    _apply_table_style(ws, header_row, ws.max_row, 1, 6)
    _autosize_columns(ws)

    ws2 = wb.create_sheet("Detalhado")
    cols = [
        "periodo",
        "arquivo",
        "base_credito_existente",
        "base_credito_geravel",
        "base_total_credito",
        "credito_pis",
        "credito_cof",
        "credito_total",
        "itens_incluidos",
        "pf_excluidos",
        "nao_vinculados_excluidos",
        "cfops_top",
        "csts_top",
        "ncms_top",
    ]
    ws2.append(cols)
    _style_header(ws2, 1, 1, len(cols))

    for r in results:
        ws2.append(
            [
                r.periodo_mmyyyy or "",
                r.path.name,
                float(q2(r.base_credito_existente)),
                float(q2(r.base_credito_geravel)),
                float(q2(r.base_total_credito)),
                float(q2(r.credito_pis)),
                float(q2(r.credito_cof)),
                float(q2(r.credito_total)),
                int(r.itens_incluidos),
                int(r.pf_excluidos),
                int(r.nao_vinculados_excluidos),
                " ".join(r.cfops_encontrados),
                " ".join(r.csts_encontrados),
                " ".join(r.ncms_encontrados),
            ]
        )

    for row in range(2, ws2.max_row + 1):
        for col in (3, 4, 5, 6, 7, 8):
            ws2.cell(row, col).number_format = _brl_number_format()

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws2.max_row}"
    _apply_table_style(ws2, 1, ws2.max_row, 1, len(cols))
    _autosize_columns(ws2)

    # ---------------------------
    # Aba 3: Itens (novo)
    # ---------------------------
    ws3 = wb.create_sheet("Itens")

    cols3 = [
        "periodo",
        "arquivo",
        "cod_part",
        "participante_nome",
        "cnpj",
        "cpf",
        "dt_doc",
        "num_doc",
        "chave",
        "cod_item",
        "descr_item",
        "ncm",
        "cfop",
        "vl_item",
        "vl_desc",
        "vl_icms",
        "vl_ipi",
        "cst_pis_original",
        "cst_pis_simulado",
        "cod_cta",
        "conta_nome",
        "conta_classif",
    ]
    # Debug rápido: ver como está vindo do result
    amostras = 0
    for r in results:
        for it in (r.itens_rows or []):
            print("[DBG IT]", "cod_part=", repr(it.cod_part), "cpf=", repr(it.participante_cpf), "nome=",
                  repr(it.participante_nome), "descr=", repr(it.descr_item))
            amostras += 1
            if amostras >= 5:
                break
        if amostras >= 5:
            break
    ws3.append(cols3)
    _style_header(ws3, 1, 1, len(cols3))

    # Preenche itens
    for r in results:
        for it in (r.itens_rows or []):
            ws3.append([
                it.periodo,
                it.arquivo,
                str(it.cod_part),
                it.participante_nome,
                str(it.participante_cnpj),
                str(it.participante_cpf),
                it.dt_doc,
                str(it.num_doc),
                str(it.chave),
                str(it.cod_item),
                str(it.descr_item),
                str(it.ncm),
                str(it.cfop),
                float(q2(it.vl_item)),
                float(q2(it.vl_desc)),
                float(q2(it.vl_icms)),
                float(q2(it.vl_ipi)),
                it.cst_pis_original,
                it.cst_pis_simulado,
                str(it.cod_cta),
                it.conta_nome,
                it.conta_classif,
            ])

    # Formatação monetária
    for row in range(2, ws3.max_row + 1):
        for col in (14, 15, 16, 17):  # vl_item, vl_desc, vl_icms, vl_ipi
            ws3.cell(row, col).number_format = _brl_number_format()

    # Filtro e estilo
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}{ws3.max_row}"
    _apply_table_style(ws3, 1, ws3.max_row, 1, len(cols3))
    # Colunas texto na aba Itens:
    # 3=cod_part, 5=cnpj, 6=cpf, 8=num_doc, 9=chave, 10=cod_item, 12=ncm, 13=cfop, 20=cod_cta
    text_cols = (3, 5, 6, 8, 9, 10, 12, 13, 20)
    for row in range(2, ws3.max_row + 1):
        for col in text_cols:
            ws3.cell(row, col).number_format = "@"
    _autosize_columns(ws3, max_width=80)
    wb.save(out_xlsx)


# --- LÓGICA DE PARSE ---

def dec_br(s: str | Decimal | None) -> Decimal:
    if s is None:
        return Decimal("0")
    if isinstance(s, Decimal):
        return s
    txt = str(s).strip().replace(".", "").replace(",", ".")
    try:
        return Decimal(txt)
    except Exception:
        return Decimal("0")


def q2(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def clean_line(ln: str) -> str:
    ln = (ln or "").strip().rstrip("\r\n")
    if not ln:
        return ""
    if not ln.startswith("|"):
        ln = "|" + ln
    if not ln.endswith("|"):
        ln = ln + "|"
    return ln


def split_reg_fields(ln: str) -> Tuple[str, List[str]]:
    ln = clean_line(ln)
    if not ln:
        return "", []
    parts = ln.strip("|").split("|")
    return (parts[0].upper().strip(), parts[1:]) if parts else ("", [])


def parse_periodo_from_0000(fields: List[str]) -> Optional[str]:
    for tok in fields:
        t = only_digits(tok)
        if len(t) == 8:
            return f"{t[2:4]}{t[4:8]}"
    return None


def mmYYYY_to_key(mmYYYY: str) -> int:
    s = (mmYYYY or "").strip()
    if len(s) != 6:
        return 0
    return int(s[2:6]) * 100 + int(s[0:2])

@dataclass
class FotoItemRow:
    periodo: str
    arquivo: str

    cod_part: str
    participante_nome: str
    participante_cnpj: str
    participante_cpf: str

    dt_doc: str
    num_doc: str
    chave: str

    cod_item: str
    descr_item: str
    ncm: str
    cfop: str

    vl_item: Decimal
    vl_desc: Decimal
    vl_icms: Decimal
    vl_ipi: Decimal

    cst_pis_original: str
    cst_pis_simulado: str

    cod_cta: str
    conta_nome: str
    conta_classif: str


@dataclass
class FotoArquivoResult:
    path: Path
    periodo_mmyyyy: Optional[str]

    # Bases
    base_credito_existente: Decimal
    base_credito_geravel: Decimal
    base_total_credito: Decimal

    # Créditos
    credito_pis: Decimal
    credito_cof: Decimal
    credito_total: Decimal

    # Contadores
    itens_incluidos: int
    pf_excluidos: int
    nao_vinculados_excluidos: int

    # Debug/top
    cfops_encontrados: List[str]
    csts_encontrados: List[str]
    ncms_encontrados: List[str]
    itens_rows: List["FotoItemRow"]


# --- CORE DO PROCESSAMENTO ---

def foto_recuperavel_ind_torrado(
    path: Path,
    *,
    cfops_validos: set[str],
    csts_origem: set[str],
    ncms_validos: set[str],
    bloquear_participante_desconhecido: bool = True,
    bloquear_c170_sem_c100: bool = True,
) -> FotoArquivoResult:
    """
    Base final do Foto = crédito existente (CST 50-56) + crédito gerável (CST origem, dentro do CFOP guard-rail).

    Travamentos mantidos:
    - Bloqueia PF (CPF)
    - Bloqueia participante desconhecido (por padrão)
    - Bloqueia C170 sem C100 (por padrão)
    """
    periodo_mmyyyy = None

    # COD_PART -> (CNPJ, CPF) (brutos)
    # 1ª Passada: Coleta de Cadastros (0150 / 0200 / 0500)
    part_data_map: Dict[str, Tuple[str, str]] = {}
    part_nome_map: Dict[str, str] = {}  # <-- fora do loop
    item_to_ncm: Dict[str, str] = {}
    item_to_desc: Dict[str, str] = {}  # <-- fora do loop (para descrição)
    conta_map: Dict[str, Tuple[str, str]] = {}  # <-- fora do loop

    with path.open("r", encoding="utf-8-sig", errors="ignore") as f:
        for raw in f:
            reg, fields = split_reg_fields(raw)

            if reg == "0150":
                cod = (fields[0] if len(fields) > 0 else "").strip()
                nome = (fields[1] if len(fields) > 1 else "").strip()
                cnpj, cpf = extrair_cnpj_cpf_0150(fields)

                if cod:
                    part_data_map[cod] = (cnpj, cpf)
                    part_nome_map[cod] = nome

                    cod_dig = only_digits(cod)
                    if cod_dig:
                        part_data_map[cod_dig] = (cnpj, cpf)
                        part_data_map[cod_dig.lstrip("0") or "0"] = (cnpj, cpf)
                        part_nome_map[cod_dig] = nome
                        part_nome_map[cod_dig.lstrip("0") or "0"] = nome

            elif reg == "0200":
                cod_i = (fields[0] if len(fields) > 0 else "").strip()
                desc_i = (fields[1] if len(fields) > 1 else "").strip()  # descrição do item
                ncm_i = only_digits(fields[6]) if len(fields) > 6 else ""
                if cod_i:
                    item_to_ncm[cod_i] = ncm_i
                    item_to_desc[cod_i] = desc_i

            elif reg == "0500":
                cod_cta = (fields[4] if len(fields) > 4 else "").strip()
                nome_cta = (fields[5] if len(fields) > 5 else "").strip()
                classif = (fields[6] if len(fields) > 6 else "").strip()
                if cod_cta:
                    conta_map[cod_cta] = (nome_cta, classif)

    # 2ª Passada: Processamento de Valores
    current_doc_is_pf: Optional[bool] = None  # None = desconhecido/sem vínculo
    current_doc_tem_c100: bool = False
    current_ind_oper: Optional[str] = None  # '0'=entrada, '1'=saída (C100)

    base_credito_existente = Decimal("0")
    base_credito_geravel = Decimal("0")

    itens_incluidos = 0
    pf_excluidos = 0
    nao_vinculados_excluidos = 0

    cfops_seen: Dict[str, int] = {}
    csts_seen: Dict[str, int] = {}
    ncms_seen: Dict[str, int] = {}

    current_cod_part = ""
    current_chv = ""
    current_cpf = ""
    current_cnpj = ""
    # excluir depois
    delta_base_pis_total = Decimal("0")
    delta_base_cof_total = Decimal("0")
    delta_desc_total = Decimal("0")
    itens_rows: List[FotoItemRow] = []

    current_doc_ignorado = False
    current_cod_sit = ""
    docs_sit_ignorado = 0  # opcional: contador

    amostras_delta = 0
    MAX_AMOSTRAS = 30  # pra não poluir o log
    ######

    with path.open("r", encoding="utf-8-sig", errors="ignore") as f:
        for raw in f:
            reg, fields = split_reg_fields(raw)

            if reg == "0000" and not periodo_mmyyyy:
                periodo_mmyyyy = parse_periodo_from_0000(fields)
                continue

            if reg == "C100":

                current_doc_tem_c100 = True
                current_num_doc = (fields[6] if len(fields) > 6 else "").strip()
                current_dt_doc = (fields[8] if len(fields) > 8 else "").strip()

                # C100: IND_OPER = fields[0]
                current_ind_oper = (fields[0] if len(fields) > 0 else "").strip()

                # C100: COD_PART = fields[2]
                cod_part = (fields[2] if len(fields) > 2 else "").strip()
                current_cod_part = cod_part

                # chave NF-e (no seu código é fields[7])
                current_chv = (fields[7] if len(fields) > 7 else "").strip()

                # C100: COD_SIT = fields[4]
                current_cod_sit = (fields[4] if len(fields) > 4 else "").strip()
                current_doc_ignorado = current_cod_sit in {"06", "07"}

                if current_doc_ignorado:
                    docs_sit_ignorado += 1  # opcional
                    print(
                        f"[FOTO][C100_SIT_SKIP] cod_sit={current_cod_sit} "
                        f"num={current_num_doc} dt={current_dt_doc} cod_part={current_cod_part} chv={current_chv}"
                    )
                    continue  # ✅ importante: não processa PF/0150 etc. para esse doc

                print("[DBG C100] COD_PART =", repr(cod_part))

                info = (
                        part_data_map.get(cod_part)
                        or part_data_map.get(only_digits(cod_part))
                        or part_data_map.get(only_digits(cod_part).lstrip("0"))
                )

                regs = [
                    v for k, v in part_data_map.items()
                    if k == cod_part or only_digits(k) == only_digits(cod_part)
                ]

                tem_cnpj = False
                tem_cpf = False

                for cnpj_raw, cpf_raw in regs:
                    if only_digits(cnpj_raw):
                        tem_cnpj = True
                        break
                    cpf = norm_cpf(cpf_raw)
                    if len(cpf) == 11 and cpf != "00000000000":
                        tem_cpf = True

                if tem_cnpj:
                    current_doc_is_pf = False
                elif tem_cpf:
                    current_doc_is_pf = True
                else:
                    current_doc_is_pf = None

                # guarda CPF/CNPJ do participante atual
                if info:
                    current_cnpj = only_digits(info[0])
                    current_cpf = norm_cpf(info[1])
                else:
                    current_cnpj = ""
                    current_cpf = ""

                continue
            if reg == "C170":
                # ✅ Se o C100 atual é complementar/cancelado, ignora todos os itens
                if current_doc_ignorado:
                    nao_vinculados_excluidos += 1  # ou crie um contador próprio
                    continue
                # 1) C170 antes de qualquer C100 = estrutura ruim -> bloqueia se configurado
                if bloquear_c170_sem_c100 and not current_doc_tem_c100:
                    nao_vinculados_excluidos += 1
                    continue

                # 1b) Espelho do motor/export: só ENTRADAS (C100 IND_OPER='0')
                if (current_ind_oper or "").strip() != "0":
                    nao_vinculados_excluidos += 1
                    continue

                # 2) Se doc PF -> bloqueia sempre
                if current_doc_is_pf is True:
                    pf_excluidos += 1
                    print(
                        f"[FOTO][PF_SKIP] cpf={current_cpf} cnpj={current_cnpj} "
                        f"cod_part={current_cod_part} chv={current_chv}"
                    )
                    continue

                # 3) Se doc desconhecido (sem 0150) -> bloqueia (espelho do motor)
                if bloquear_participante_desconhecido and current_doc_is_pf is None:
                    nao_vinculados_excluidos += 1
                    print(
                        f"[FOTO][SEM_0150_SKIP] cpf={current_cpf} cnpj={current_cnpj} "
                        f"cod_part={current_cod_part} chv={current_chv}"
                    )
                    continue

                # --- agora sim: extrair campos ---
                vl_item = dec_br(fields[5]) if len(fields) > 5 else Decimal("0")
                cfop = (fields[9] if len(fields) > 9 else "").strip()
                cst_pis = (fields[23] if len(fields) > 23 else "").strip()
                cod_item = (fields[1] if len(fields) > 1 else "").strip()

                # contábil: COD_CTA vem no final do C170 (no seu exemplo: 55)
                cod_cta_item = (fields[-1] if fields else "").strip()
                cod_cta_final = cod_cta_item or ""
                conta_nome, conta_classif = conta_map.get(cod_cta_final, ("", ""))

                # NCM + descrição (fallback 0200)
                ncm = item_to_ncm.get(cod_item, "") if cod_item else ""
                descr_item_c170 = (fields[2] if len(fields) > 2 else "").strip()
                descr_item = descr_item_c170 or item_to_desc.get(cod_item, "")

                # participante nome (0150)
                participante_nome = (
                        part_nome_map.get(current_cod_part)
                        or part_nome_map.get(only_digits(current_cod_part))
                        or part_nome_map.get(only_digits(current_cod_part).lstrip("0"))
                        or ""
                )

                # guard-rail CFOP + valor
                if cfop not in cfops_validos or vl_item <= 0:
                    continue

                # gate NCM (igual você já tinha)
                if ncms_validos:
                    if ncm not in ncms_validos:
                        continue
                else:
                    if not ncm.startswith("09"):
                        continue

                if ncm:
                    ncms_seen[ncm] = ncms_seen.get(ncm, 0) + 1

                # ✅ (A) já é crédito (CST 50-56)
                if cst_pis in CST_CREDITO:
                    base_credito_existente += vl_item
                    itens_incluidos += 1
                    cfops_seen[cfop] = cfops_seen.get(cfop, 0) + 1
                    csts_seen[cst_pis] = csts_seen.get(cst_pis, 0) + 1

                    itens_rows.append(
                        FotoItemRow(
                            periodo=periodo_mmyyyy or "",
                            arquivo=path.name,

                            cod_part=current_cod_part,
                            participante_nome=participante_nome,
                            participante_cnpj=current_cnpj,
                            participante_cpf=current_cpf,

                            dt_doc=current_dt_doc,
                            num_doc=current_num_doc,
                            chave=current_chv,

                            cod_item=cod_item,
                            descr_item=descr_item,
                            ncm=ncm,
                            cfop=cfop,

                            vl_item=vl_item,
                            vl_desc=dec_br(fields[6]) if len(fields) > 6 else Decimal("0"),
                            vl_icms=dec_br(fields[13]) if len(fields) > 13 else Decimal("0"),
                            vl_ipi=dec_br(fields[17]) if len(fields) > 17 else Decimal("0"),

                            cst_pis_original=cst_pis,
                            cst_pis_simulado=cst_pis,

                            cod_cta=cod_cta_final,
                            conta_nome=conta_nome,
                            conta_classif=conta_classif,
                        )
                    )
                    continue

                # ✅ (B) vai virar crédito com as alterações (CST origem)
                if cst_pis in csts_origem:
                    base_credito_geravel += vl_item
                    itens_incluidos += 1
                    cfops_seen[cfop] = cfops_seen.get(cfop, 0) + 1
                    csts_seen[cst_pis] = csts_seen.get(cst_pis, 0) + 1

                    itens_rows.append(
                        FotoItemRow(
                            periodo=periodo_mmyyyy or "",
                            arquivo=path.name,

                            cod_part=current_cod_part,
                            participante_nome=participante_nome,
                            participante_cnpj=current_cnpj,
                            participante_cpf=current_cpf,

                            dt_doc=current_dt_doc,
                            num_doc=current_num_doc,
                            chave=current_chv,

                            cod_item=cod_item,
                            descr_item=descr_item,
                            ncm=ncm,
                            cfop=cfop,

                            vl_item=vl_item,
                            vl_desc=dec_br(fields[6]) if len(fields) > 6 else Decimal("0"),
                            vl_icms=dec_br(fields[13]) if len(fields) > 13 else Decimal("0"),
                            vl_ipi=dec_br(fields[17]) if len(fields) > 17 else Decimal("0"),

                            cst_pis_original=cst_pis,
                            cst_pis_simulado="51",  # simulação do alvo

                            cod_cta=cod_cta_final,
                            conta_nome=conta_nome,
                            conta_classif=conta_classif,
                        )
                    )
                    continue

                if vl_item == Decimal("80"):
                    print("ITEM80", cod_item, ncm, cst_pis, cfop, current_chv)

                # demais CSTs: fora do escopo do Foto
                continue


    base_credito_existente = q2(base_credito_existente)
    base_credito_geravel = q2(base_credito_geravel)
    base_total_credito = q2(base_credito_existente + base_credito_geravel)

    credito_pis = q2(base_total_credito * ALIQUOTA_PIS)
    credito_cof = q2(base_total_credito * ALIQUOTA_COF)
    credito_total = q2(credito_pis + credito_cof)

    print("\n[FOTO][RESUMO_DELTAS]")
    print("  delta_base_pis_total (VL_ITEM - VL_BC_PIS) =", fmt_br(q2(delta_base_pis_total)))
    print("  delta_base_cof_total (VL_ITEM - VL_BC_COF) =", fmt_br(q2(delta_base_cof_total)))
    print("  total_vl_desc (soma VL_DESC)               =", fmt_br(q2(delta_desc_total)))

    # Se você tem o crédito do M (export/PVA) em mãos, você consegue checar a base equivalente:
    # base_equiv_pis = q2((credito_pis - credito_pis_m) / ALIQUOTA_PIS)
    print(f"[FOTO] itens_rows capturados = {len(itens_rows)}")
    return FotoArquivoResult(
        path=path,
        periodo_mmyyyy=periodo_mmyyyy,

        base_credito_existente=base_credito_existente,
        base_credito_geravel=base_credito_geravel,
        base_total_credito=base_total_credito,

        credito_pis=credito_pis,
        credito_cof=credito_cof,
        credito_total=credito_total,

        itens_incluidos=itens_incluidos,
        pf_excluidos=pf_excluidos,
        nao_vinculados_excluidos=nao_vinculados_excluidos,

        cfops_encontrados=sorted(cfops_seen.keys()),
        csts_encontrados=sorted(csts_seen.keys()),
        ncms_encontrados=sorted(ncms_seen.keys()),
        itens_rows=itens_rows,
    )


def fmt_br(d: Decimal) -> str:
    return f"{d:,.2f}".replace(",", "v").replace(".", ",").replace("v", ".")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("pasta")
    ap.add_argument("--ncm", action="append", default=[])
    ap.add_argument("--out", default="")

    # ✅ flags opcionais, mas default = seguro (bloqueia desconhecido)
    ap.add_argument("--permitir-desconhecido", action="store_true", help="Não bloquear quando COD_PART não for encontrado no 0150")
    ap.add_argument("--permitir-c170-sem-c100", action="store_true", help="Não bloquear C170 que apareça antes de qualquer C100")

    args = ap.parse_args()

    pasta = Path(args.pasta)
    if not pasta.is_dir():
        return 1

    ncms_validos = set(only_digits(x) for x in args.ncm if x)

    results = [
        foto_recuperavel_ind_torrado(
            p,
            cfops_validos=set(CFOPS_TORRADO_PADRAO),
            csts_origem=set(CSTS_ORIGEM_PADRAO),
            ncms_validos=ncms_validos,
            bloquear_participante_desconhecido=not args.permitir_desconhecido,
            bloquear_c170_sem_c100=not args.permitir_c170_sem_c100,
        )
        for p in sorted(pasta.glob("*.txt"))
    ]
    results.sort(key=lambda x: mmYYYY_to_key(x.periodo_mmyyyy or ""))

    total_base_exist = sum((r.base_credito_existente for r in results), Decimal("0"))
    total_base_ger = sum((r.base_credito_geravel for r in results), Decimal("0"))
    total_base_total = sum((r.base_total_credito for r in results), Decimal("0"))
    total_cred = sum((r.credito_total for r in results), Decimal("0"))

    print(f"\nTOTAL base_credito_existente=R$ {fmt_br(q2(total_base_exist))}")
    print(f"TOTAL base_credito_geravel=R$    {fmt_br(q2(total_base_ger))}")
    print(f"TOTAL base_total_credito=R$      {fmt_br(q2(total_base_total))}")
    print(f"TOTAL crédito=R$                 {fmt_br(q2(total_cred))}")

    if args.out and args.out.endswith(".xlsx"):
        salvar_xlsx_visual(
            Path(args.out),
            pasta=pasta,
            ncms_validos=ncms_validos,
            results=results,
            total_base_credito_existente=total_base_exist,
            total_base_credito_geravel=total_base_ger,
            total_base_total_credito=total_base_total,
            total_cred=total_cred,
        )
    return 0


if __name__ == "__main__":
    main()
